from django.shortcuts import render,redirect,reverse
from . import forms,models
from django.db.models import Sum
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required,user_passes_test
from django.conf import settings
from datetime import date, timedelta
from exam import models as QMODEL
from teacher import models as TMODEL
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.core.mail import send_mail

def student_login_view(request):
    from django.contrib.auth.forms import AuthenticationForm
    form = AuthenticationForm()
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                try:
                    student = models.Student.objects.get(user=user)
                    if student.status:
                        login(request, user)
                        return redirect('student_dashboard')
                    else:
                        return render(request, 'student/student_wait_for_approval.html')
                except models.Student.DoesNotExist:
                    messages.error(request, 'Invalid credentials.')
                    return render(request, 'student/studentlogin.html', {'form': form})
            else:
                messages.error(request, 'Invalid credentials.')
                return render(request, 'student/studentlogin.html', {'form': form})
        else:
            return render(request, 'student/studentlogin.html', {'form': form})
    return render(request, 'student/studentlogin.html', {'form': form})
from django.db.models import Max

def generate_enrollment_number():
    """Generate a unique enrollment number for students"""
    # Get the maximum enrollment number more efficiently
    max_enrollment = models.Student.objects.filter(enrollment_number__isnull=False).aggregate(
        max_enrollment=Max('enrollment_number')
    )['max_enrollment']

    if max_enrollment:
        try:
            # Extract number from format STU0001
            num_str = max_enrollment[3:]
            num = int(num_str) + 1
            new_enrollment = f"STU{num:04d}"
        except (ValueError, IndexError):
            new_enrollment = "STU0001"
    else:
        new_enrollment = "STU0001"

    # Ensure uniqueness by incrementing if needed
    while models.Student.objects.filter(enrollment_number=new_enrollment).exists():
        try:
            num_str = new_enrollment[3:]
            num = int(num_str) + 1
            new_enrollment = f"STU{num:04d}"
        except (ValueError, IndexError):
            new_enrollment = "STU0001"

    return new_enrollment


#for showing signup/login button for student
def studentclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse('student_dashboard'))  # Updated redirect
    userForm = forms.StudentUserForm()
    studentForm = forms.StudentForm()
    mydict = {'userForm': userForm, 'studentForm': studentForm}
    return render(request,'student/studentclick.html', context=mydict)

from studentprofile.models import StudentProfileRequest

def registration_view(request):
    userForm = forms.StudentUserForm()
    studentForm = forms.StudentForm()
    mydict = {'userForm': userForm, 'studentForm': studentForm}

    if request.method == 'POST':
        if 'register' in request.POST:
            userForm = forms.StudentUserForm(request.POST, prefix='userForm')
            studentForm = forms.StudentForm(request.POST, request.FILES, prefix='studentForm')
            if userForm.is_valid() and studentForm.is_valid():
                user = userForm.save()
                user.set_password(user.password)
                user.save()
                student = studentForm.save(commit=False)
                student.user = user
                student.status = False  # Require approval
                student.enrollment_number = generate_enrollment_number()  # Generate unique enrollment number
                student.save()
                my_student_group = Group.objects.get_or_create(name='STUDENT')
                my_student_group[0].user_set.add(user)
                # Send registration confirmation email to student
                subject = 'Registration Successful - LJ University'
                message = f'Dear {user.first_name},\n\nYour student account has been created successfully. Your application is pending approval from the administrator and teacher. You will receive an email once your account is approved.\n\nBest regards,\nLJ University Administration'
                from_email = settings.EMAIL_HOST_USER
                recipient_list = [user.email]
                send_mail(subject, message, from_email, recipient_list, fail_silently=False)

                # Send notification email to admins and teachers
                subject_admin = 'New Student Registration Pending Approval - LJ University'
                message_admin = f'Dear Administrator/Teacher,\n\nA new student {user.first_name} {user.last_name} ({user.username}) has registered and is pending approval.\n\nPlease log in to the system to review and approve the student.\n\nBest regards,\nLJ University System'
                from_email = settings.EMAIL_HOST_USER
                # Get all admin emails (superusers)
                admin_emails = list(User.objects.filter(is_superuser=True).values_list('email', flat=True))
                # Get all teacher emails
                teacher_emails = list(TMODEL.Teacher.objects.filter(status=True).values_list('user__email', flat=True))
                recipient_list_admin = admin_emails + teacher_emails
                if recipient_list_admin:
                    send_mail(subject_admin, message_admin, from_email, recipient_list_admin, fail_silently=False)
                # Redirect to waiting for approval page
                return render(request, 'student/student_wait_for_approval.html')
            else:
                mydict['userForm'] = userForm
                mydict['studentForm'] = studentForm


    return render(request, 'student/registration.html', context=mydict)

def student_signup_view(request):
    userForm=forms.StudentUserForm()
    studentForm=forms.StudentForm()
    mydict={'userForm':userForm,'studentForm':studentForm}
    if request.method=='POST':
        userForm=forms.StudentUserForm(request.POST)
        studentForm=forms.StudentForm(request.POST,request.FILES)
        if userForm.is_valid() and studentForm.is_valid():
            user=userForm.save()
            user.set_password(user.password)
            user.save()
            student=studentForm.save(commit=False)
            student.user=user
            student.status = False  # Require approval
            student.enrollment_number = generate_enrollment_number()  # Generate unique enrollment number
            student.save()
            my_student_group = Group.objects.get_or_create(name='STUDENT')
            my_student_group[0].user_set.add(user)
            # Redirect to waiting for approval page
            return render(request, 'student/student_wait_for_approval.html')
        else:
            messages.error(request, 'Please correct the errors below.')
    return render(request,'student/registration.html',context=mydict)

def is_student(user):
    return user.groups.filter(name='STUDENT').exists()

def is_student_approved(user):
    """Check if student is approved (has status=True)"""
    if is_student(user):
        try:
            student = models.Student.objects.get(user=user)
            return student.status
        except models.Student.DoesNotExist:
            return False
    return False

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_dashboard_view(request):
    student = models.Student.objects.get(user=request.user)
    if not student.status:
        return render(request, 'student/student_wait_for_approval.html')

    # Return the dashboard with personalized student data
    dict={
        'total_course':QMODEL.Course.objects.all().count(),
        'total_question':QMODEL.Question.objects.all().count(),
        'student': student,
    }
    return render(request,'student/student_dashboard.html',context=dict)

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_exam_view(request):
    # Check if student is approved
    student = models.Student.objects.get(user=request.user)
    if not student.status:
        return render(request, 'student/student_wait_for_approval.html')
    
    courses=QMODEL.Course.objects.all()
    return render(request,'student/student_exam.html',{'courses':courses})

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def take_exam_view(request,pk):
    # Check if student is approved
    student = models.Student.objects.get(user=request.user)
    if not student.status:
        return render(request, 'student/student_wait_for_approval.html')

    course=QMODEL.Course.objects.get(id=pk)
    total_questions=QMODEL.Question.objects.all().filter(course=course).count()
    questions=QMODEL.Question.objects.all().filter(course=course)
    total_marks=0
    for q in questions:
        total_marks=total_marks + q.marks

    return render(request,'student/take_exam.html',{'course':course,'total_questions':total_questions,'total_marks':total_marks})

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def start_exam_view(request,pk):
    course=QMODEL.Course.objects.get(id=pk)
    student = models.Student.objects.get(user_id=request.user.id)
    if QMODEL.Result.objects.filter(exam=course, student=student).exists():
        return HttpResponseRedirect(reverse('view_result'))
    questions=QMODEL.Question.objects.all().filter(course=course)
    if request.method=='POST':
        pass
    response= render(request,'student/start_exam.html',{'course':course,'questions':questions})
    response.set_cookie('course_id',course.id)
    return response


@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def calculate_marks_view(request):
    if request.method == 'POST':
        # Get course_id from POST data instead of cookie
        course_id = request.POST.get('course_id')
        if course_id:
            try:
                course = QMODEL.Course.objects.get(id=course_id)
                total_marks = 0
                questions = QMODEL.Question.objects.all().filter(course=course)

                for i in range(len(questions)):
                    selected_ans = request.POST.get(str(i+1))
                    actual_answer = questions[i].answer
                    if selected_ans == actual_answer:
                        total_marks = total_marks + questions[i].marks

                student = models.Student.objects.get(user_id=request.user.id)
                result = QMODEL.Result()
                result.marks = total_marks
                result.exam = course
                result.student = student
                result.save()

                return HttpResponseRedirect(reverse('view_result'))
            except QMODEL.Course.DoesNotExist:
                pass
    return HttpResponseRedirect(reverse('student_dashboard'))



@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def view_result_view(request):
    courses=QMODEL.Course.objects.all()
    return render(request,'student/view_result.html',{'courses':courses})
    

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def check_marks_view(request,pk):
    course=QMODEL.Course.objects.get(id=pk)
    student = models.Student.objects.get(user_id=request.user.id)
    results= QMODEL.Result.objects.all().filter(exam=course).filter(student=student)
    return render(request,'student/check_marks.html',{'results':results})

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_marks_view(request):
    courses=QMODEL.Course.objects.all()
    return render(request,'student/student_marks.html',{'courses':courses})

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_profile_view(request):
    student = models.Student.objects.get(user=request.user)

    if request.method == 'POST':
        enrollment = request.POST.get('enrollment_number')
        if enrollment:
            student.enrollment_number = enrollment
            student.save()
            messages.success(request, 'Enrollment number updated successfully. You can now access the paper system.')

            # Send email notification to student
            subject = 'Enrollment Number Updated'
            message = f'Dear {request.user.first_name},\n\nYour enrollment number has been successfully updated to {enrollment}. You can now access the paper system.\n\nBest regards,\nLJ University'
            from_email = None  # Uses DEFAULT_FROM_EMAIL if set, else EMAIL_HOST_USER
            recipient_list = [request.user.email]
            send_mail(subject, message, from_email, recipient_list, fail_silently=False)

            return redirect('student_dashboard')
        else:
            messages.error(request, 'Please enter a valid enrollment number.')

    # Paper system data - optimized to avoid slow Excel processing on every load
    from studentprofile.models import Paper, PaperRequest, StudentPaper

    # Only load paper data if student has enrollment number
    if student.enrollment_number:
        # Show papers where the student is selected or all papers if no selection
        papers = Paper.objects.filter(selected_students=student) | Paper.objects.filter(selected_students__isnull=True)
        student_papers = []

        # Limit Excel processing to avoid slow loading - only process first 5 papers
        for paper in papers[:5]:  # Limit to prevent slow loading
            try:
                from openpyxl import load_workbook
                wb = load_workbook(paper.excel_file.path)
                sheet = wb.active
                data = []
                # Limit rows to prevent memory issues
                for row_num, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_num >= 100:  # Limit to first 100 rows
                        break
                    data.append(row)

                # Filter data to show only the row where enrollment matches student's enrollment
                filtered_data = []
                for row in data:
                    if len(row) > 1 and row[1] == student.enrollment_number:
                        filtered_data.append(row)
                paper.excel_data = filtered_data
            except Exception as e:
                paper.excel_data = None
            student_papers.append(paper)

        requests = PaperRequest.objects.filter(student=student)
        approved_paper_ids = set(req.paper.id for req in requests if req.status == 'approved')

        # Add download_file attribute to approved requests for proper paper file access
        for req in requests:
            if req.status == 'approved':
                try:
                    student_paper = StudentPaper.objects.get(paper=req.paper, student=student)
                    req.download_file = student_paper.paper_file
                except StudentPaper.DoesNotExist:
                    req.download_file = req.paper.paper_file
    else:
        # If no enrollment number, don't load heavy paper data
        student_papers = []
        requests = []
        approved_paper_ids = set()

    return render(request, 'student/student_profile.html', {'student': student, 'papers': student_papers, 'requests': requests, 'approved_paper_ids': approved_paper_ids})
