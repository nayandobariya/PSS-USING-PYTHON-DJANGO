from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseNotFound
from .models import StudentProfileRequest, Paper, PaperRequest, Feedback, StudentPaper
from .forms import StudentProfileRequestForm, PaperRequestForm, FeedbackForm, PaperUploadForm
from student.models import Student
from teacher.models import Teacher
from exam.models import Course
from django.contrib.auth.models import Group

def is_student(user):
    return user.groups.filter(name='STUDENT').exists()

def is_teacher(user):
    return user.groups.filter(name='TEACHER').exists()

def is_admin(user):
    return user.is_superuser

@login_required
@user_passes_test(is_student)
def student_profile_view(request):
    from openpyxl import load_workbook
    student = get_object_or_404(Student, user=request.user)
    # Show papers where the student is selected or all papers if no selection
    papers = Paper.objects.filter(selected_students=student) | Paper.objects.filter(selected_students__isnull=True)
    student_papers = []
    for paper in papers:
        try:
            wb = load_workbook(paper.excel_file.path)
            sheet = wb.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            # Filter data to show only the row where enrollment matches student's enrollment
            filtered_data = []
            for row in data:
                if len(row) > 1 and row[1] == student.enrollment_number:
                    filtered_data.append(row)
            paper.excel_data = filtered_data
        except Exception as e:
            paper.excel_data = None

        # Check if student has individual paper file
        try:
            student_paper = StudentPaper.objects.get(paper=paper, student=student)
            paper.download_file = student_paper.paper_file
        except StudentPaper.DoesNotExist:
            paper.download_file = paper.paper_file

        student_papers.append(paper)
    # Show individual student papers
    individual_papers = StudentPaper.objects.filter(student=student)
    # Keep request logic for other features
    requests = PaperRequest.objects.filter(student=student)
    approved_paper_ids = set(req.paper.id for req in requests if req.status == 'approved')
    requested_paper_ids = set(req.paper.id for req in requests)
    # Get paper requests that have feedback given
    feedback_given_requests = set(Feedback.objects.filter(paper_request__in=requests).values_list('paper_request_id', flat=True))

    # Attach download_file to each request for proper file access
    for req in requests:
        try:
            student_paper = StudentPaper.objects.get(paper=req.paper, student=student)
            req.download_file = student_paper.paper_file
        except StudentPaper.DoesNotExist:
            req.download_file = req.paper.paper_file
    return render(request, 'studentprofile/student_profile.html', {
        'student': student,
        'papers': student_papers,
        'individual_papers': individual_papers,
        'requests': requests,
        'approved_paper_ids': approved_paper_ids,
        'paper_requests': requests,
        'requested_paper_ids': requested_paper_ids,
        'feedback_given_requests': feedback_given_requests,
    })

@login_required
@user_passes_test(is_student)
def request_form_view(request):
    if request.method == 'POST':
        form = StudentProfileRequestForm(request.POST)
        if form.is_valid():
            req = form.save(commit=False)
            req.student = get_object_or_404(Student, user=request.user)
            req.save()
            messages.success(request, 'Request submitted successfully.')
            return redirect('studentprofile:student_profile')
    else:
        form = StudentProfileRequestForm()
    return render(request, 'studentprofile/request_form.html', {'form': form})

@login_required
@user_passes_test(is_student)
def raise_request_view(request, paper_id):
    paper = get_object_or_404(Paper, id=paper_id)
    student = get_object_or_404(Student, user=request.user)
    if not student.status:
        messages.error(request, 'Your account is not approved yet. Please wait for admin approval.')
        return redirect('student_dashboard')
    if not PaperRequest.objects.filter(student=student, paper=paper).exists():
        PaperRequest.objects.create(student=student, paper=paper)
        messages.success(request, 'Request raised successfully.')
    else:
        messages.warning(request, 'Request already exists.')
    return redirect('studentprofile:student_profile')

@login_required
@user_passes_test(is_student)
def give_feedback_view(request, request_id):
    paper_request = get_object_or_404(PaperRequest, id=request_id, student__user=request.user)
    student = get_object_or_404(Student, user=request.user)
    if not student.status:
        messages.error(request, 'Your account is not approved yet. Please wait for admin approval.')
        return redirect('student_dashboard')
    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.paper_request = paper_request
            feedback.save()
            messages.success(request, 'Feedback submitted.')
            return redirect('studentprofile:student_profile')
    else:
        form = FeedbackForm()
    return render(request, 'studentprofile/give_feedback.html', {'form': form, 'paper_request': paper_request})

@login_required
@user_passes_test(is_teacher)
def teacher_upload_view(request):
    if request.method == 'POST':
        if 'generate' in request.POST:
            course_id = request.POST.get('course')
            selected_student_ids = request.POST.getlist('selected_students')
            if not selected_student_ids:
                messages.error(request, 'Please select at least one student.')
                form = PaperUploadForm()
                courses = Course.objects.all()
                students = Student.objects.filter(status=True)
                return render(request, 'studentprofile/teacher_upload.html', {'courses': courses, 'students': students, 'form': form})
            course = get_object_or_404(Course, id=course_id)
            students = Student.objects.filter(id__in=selected_student_ids, status=True)
            from openpyxl import Workbook
            from django.http import HttpResponse
            wb = Workbook()
            ws = wb.active
            ws.title = f"{course.course_name} Papers"
            headers = ['Name', 'Enrollment', 'Marks', 'Feedback']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            for row_num, student in enumerate(students, 2):
                ws.cell(row=row_num, column=1, value=student.get_name)
                ws.cell(row=row_num, column=2, value=student.enrollment_number)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{course.course_name}_papers_template.xlsx"'
            wb.save(response)
            return response
        else:
            form = PaperUploadForm(request.POST, request.FILES)
            if form.is_valid():
                selected_students = form.cleaned_data['selected_students']
                paper = form.save(commit=False)
                paper.teacher = get_object_or_404(Teacher, user=request.user)
                paper.save()
                paper.selected_students.set(selected_students)
                # Create individual papers for each selected student
                for student in selected_students:
                    file_key = f'student_files[{student.id}]'
                    if file_key in request.FILES:
                        StudentPaper.objects.create(paper=paper, student=student, paper_file=request.FILES[file_key])
                messages.success(request, f'Papers uploaded successfully for {len(selected_students)} students.')
                return redirect('studentprofile:teacher_dashboard')
            else:
                messages.error(request, 'Form is invalid.')
    else:
        form = PaperUploadForm()
    courses = Course.objects.all()
    students = Student.objects.filter(status=True)
    return render(request, 'studentprofile/teacher_upload.html', {'courses': courses, 'students': students, 'form': form})

@login_required
@user_passes_test(is_teacher)
def teacher_dashboard_view(request):
    teacher = get_object_or_404(Teacher, user=request.user)
    papers = Paper.objects.filter(teacher=teacher)
    student_papers = StudentPaper.objects.filter(paper__teacher=teacher)
    requests = PaperRequest.objects.filter(paper__teacher=teacher)
    feedbacks = Feedback.objects.filter(paper_request__paper__teacher=teacher)
    return render(request, 'studentprofile/teacher_dashboard.html', {
        'papers': papers,
        'student_papers': student_papers,
        'requests': requests,
        'feedbacks': feedbacks,
    })

@login_required
@user_passes_test(is_admin)
def admin_dashboard_view(request):
    requests = StudentProfileRequest.objects.all()
    paper_requests = PaperRequest.objects.all()
    feedbacks = Feedback.objects.all()
    return render(request, 'studentprofile/admin_dashboard.html', {
        'requests': requests,
        'paper_requests': paper_requests,
        'feedbacks': feedbacks,
    })

@login_required
@user_passes_test(is_admin)
def approve_request_view(request, request_id):
    req = get_object_or_404(StudentProfileRequest, id=request_id)
    req.status = 'approved'
    req.save()
    # Update the student's status to approved
    student = req.student
    student.status = True
    student.save()
    # Send approval email
    subject = 'Account Approved - LJ University'
    message = f'Dear {student.user.first_name},\n\nYour student account has been approved. You can now log in and access the system.\n\nBest regards,\nLJ University Administration'
    from_email = settings.EMAIL_HOST_USER
    recipient_list = [student.user.email]
    send_mail(subject, message, from_email, recipient_list, fail_silently=False)
    messages.success(request, 'Request approved.')
    return redirect('studentprofile:admin_dashboard')

@login_required
@user_passes_test(is_admin)
def approve_paper_request_view(request, pr_id):
    pr = get_object_or_404(PaperRequest, id=pr_id)
    pr.status = 'approved'
    pr.save()
    messages.success(request, 'Paper request approved.')
    return redirect('studentprofile:admin_dashboard')
